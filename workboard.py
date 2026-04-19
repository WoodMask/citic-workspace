# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, date
import os

EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '模型推理服务.xlsx')
HTML_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), '工作看板.html')

STATUS_COLORS = {
    '已完成': '#10b981',
    '进行中': '#3b82f6', 
    '未启动': '#6b7280',
}

STATUS_ICONS = {
    '已完成': '✅',
    '进行中': '🔄',
    '未启动': '⏳',
}

def migrate_task_ids(ws):
    header_id = ws.cell(row=1, column=1).value
    if header_id is None or '编号' not in str(header_id):
        ws.insert_cols(1)
        ws.cell(row=1, column=1).value = '任务编号'
        ws.cell(row=1, column=2).value = '任务'
        ws.cell(row=1, column=3).value = '任务状态'
        ws.cell(row=1, column=4).value = '开始日期'
        ws.cell(row=1, column=5).value = '完成日期'
        ws.cell(row=1, column=6).value = '交付物'
        ws.cell(row=1, column=7).value = '备注'
    
    task_id = 1
    for row_idx in range(2, ws.max_row + 1):
        status = ws.cell(row=row_idx, column=3).value
        if status is not None and str(status).strip() in ['已完成', '进行中', '未启动']:
            existing_id = ws.cell(row=row_idx, column=1).value
            if existing_id is None or str(existing_id).strip() == '':
                ws.cell(row=row_idx, column=1).value = f'{task_id:03d}'
                task_id += 1
            else:
                try:
                    task_id = max(task_id, int(str(existing_id).strip()) + 1)
                except:
                    ws.cell(row=row_idx, column=1).value = f'{task_id:03d}'
                    task_id += 1

def read_excel_data():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    migrate_task_ids(ws)
    wb.save(EXCEL_FILE)
    
    categories = []
    current_category = None
    is_milestone_section = False
    tasks = []
    milestones = []
    
    for row_idx in range(2, ws.max_row + 1):
        task_id = ws.cell(row=row_idx, column=1).value
        task_name = ws.cell(row=row_idx, column=2).value
        status = ws.cell(row=row_idx, column=3).value
        start_date = ws.cell(row=row_idx, column=4).value
        end_date = ws.cell(row=row_idx, column=5).value
        deliverable = ws.cell(row=row_idx, column=6).value
        note = ws.cell(row=row_idx, column=7).value
        
        if task_name is None or str(task_name).strip() == '':
            continue
        
        task_str = str(task_name).strip()
        
        if task_str.startswith(('一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '里程碑')):
            if current_category and tasks:
                categories.append({'name': current_category, 'tasks': tasks})
                tasks = []
            current_category = task_str
            is_milestone_section = task_str == '里程碑'
        elif status is not None and str(status).strip() in ['已完成', '进行中', '未启动']:
            task_data = {
                'row': row_idx,
                'id': str(task_id).strip() if task_id else f'{row_idx:03d}',
                'name': task_str,
                'status': str(status).strip(),
                'start_date': start_date,
                'end_date': end_date,
                'deliverable': deliverable if deliverable else '',
                'note': note if note else '',
            }
            if is_milestone_section:
                milestones.append(task_data)
            else:
                tasks.append(task_data)
    
    if current_category and tasks:
        categories.append({'name': current_category, 'tasks': tasks})
    
    wb.close()
    
    return categories, milestones

def calculate_date_status(start_date, end_date, task_status):
    today = date.today()
    
    if task_status == '已完成':
        return 'normal', ''
    
    if start_date and task_status == '未启动':
        if isinstance(start_date, datetime):
            start = start_date.date()
        elif isinstance(start_date, date):
            start = start_date
        else:
            start = None
        
        if start and start <= today:
            days_passed = (today - start).days
            if days_passed == 0:
                return 'should_start', '今日应启动'
            else:
                return 'should_start', f'应启动{days_passed}天'
    
    if end_date is None:
        return 'normal', ''
    
    if isinstance(end_date, datetime):
        end = end_date.date()
    elif isinstance(end_date, date):
        end = end_date
    else:
        return 'normal', ''
    
    days_diff = (end - today).days
    
    if days_diff < 0 and task_status == '进行中':
        return 'overdue', f'超期{abs(days_diff)}天'
    if days_diff <= 3 and days_diff >= 0 and task_status == '进行中':
        return 'warning', f'{days_diff}天后到期'
    if days_diff < 0 and task_status == '未启动':
        return 'soon', f'超期{abs(days_diff)}天'
    return 'normal', ''

def format_date(d):
    if d is None:
        return '未设置'
    if isinstance(d, datetime):
        return d.strftime('%m/%d')
    if isinstance(d, date):
        return d.strftime('%m/%d')
    return str(d)

CSS_STYLES = '''
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { 
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        .container { max-width: 1800px; margin: 0 auto; }
        
        .header {
            background: white;
            border-radius: 16px;
            padding: 24px;
            margin-bottom: 20px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }
        .header-top {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
        }
        .header h1 {
            font-size: 28px;
            color: #1f2937;
            display: flex;
            align-items: center;
            gap: 12px;
        }
        .header-date {
            font-size: 16px;
            color: #6b7280;
        }
        
        .progress-section {
            margin-bottom: 20px;
        }
        .progress-label {
            display: flex;
            justify-content: space-between;
            margin-bottom: 8px;
            font-size: 14px;
            color: #4b5563;
        }
        .progress-bar {
            height: 24px;
            background: #e5e7eb;
            border-radius: 12px;
            overflow: hidden;
        }
        .progress-fill {
            height: 100%;
            background: linear-gradient(90deg, #10b981, #34d399);
            border-radius: 12px;
            transition: width 0.5s ease;
            display: flex;
            align-items: center;
            justify-content: flex-end;
            padding-right: 12px;
            color: white;
            font-weight: bold;
            font-size: 12px;
        }
        
        .stats {
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 16px;
            margin-top: 20px;
        }
        .stat-card {
            background: #f9fafb;
            border-radius: 12px;
            padding: 16px;
            text-align: center;
        }
        .stat-value {
            font-size: 32px;
            font-weight: bold;
            color: #1f2937;
        }
        .stat-label {
            font-size: 14px;
            color: #6b7280;
            margin-top: 4px;
        }
        .stat-card.completed .stat-value { color: #10b981; }
        .stat-card.progress .stat-value { color: #3b82f6; }
        .stat-card.pending .stat-value { color: #6b7280; }
        .stat-card.warning .stat-value { color: #ef4444; }
        
        .warnings {
            background: #fef2f2;
            border: 1px solid #fecaca;
            border-radius: 12px;
            padding: 16px;
            margin-top: 20px;
        }
        .warnings h3 {
            color: #dc2626;
            margin-bottom: 12px;
            font-size: 16px;
        }
        .warning-item {
            display: flex;
            align-items: center;
            gap: 8px;
            padding: 8px 0;
            border-bottom: 1px solid #fecaca;
            font-size: 14px;
        }
        .warning-item:last-child { border-bottom: none; }
        .warning-badge {
            background: #dc2626;
            color: white;
            padding: 2px 8px;
            border-radius: 4px;
            font-size: 12px;
        }
        .warning-badge.orange {
            background: #f59e0b;
        }
        
        .kanban {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(320px, 1fr));
            gap: 20px;
        }
        
        .category {
            background: white;
            border-radius: 16px;
            overflow: hidden;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }
        .category-header {
            padding: 16px 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }
        .category-title {
            font-size: 16px;
            font-weight: 600;
            margin-bottom: 8px;
        }
        .category-progress {
            display: flex;
            align-items: center;
            gap: 8px;
            font-size: 12px;
        }
        .category-progress-bar {
            flex: 1;
            height: 6px;
            background: rgba(255,255,255,0.3);
            border-radius: 3px;
            overflow: hidden;
        }
        .category-progress-fill {
            height: 100%;
            background: white;
            border-radius: 3px;
        }
        
        .tasks {
            padding: 12px;
            max-height: 500px;
            overflow-y: auto;
        }
        
        .task {
            background: #f9fafb;
            border-radius: 8px;
            padding: 12px;
            margin-bottom: 8px;
            border-left: 4px solid #6b7280;
            transition: transform 0.2s, box-shadow 0.2s;
        }
        .task:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        }
        .task.completed { border-left-color: #10b981; background: #f0fdf4; }
        .task.progress { border-left-color: #3b82f6; background: #eff6ff; }
        .task.pending { border-left-color: #6b7280; }
        
        .task-header {
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            margin-bottom: 8px;
        }
        .task-name {
            font-size: 14px;
            font-weight: 500;
            color: #1f2937;
            flex: 1;
            margin-right: 8px;
        }
        .task-row {
            color: #667eea;
            font-weight: bold;
        }
        .task-status {
            font-size: 12px;
            padding: 2px 8px;
            border-radius: 4px;
            color: white;
            flex-shrink: 0;
        }
        .task-status.completed { background: #10b981; }
        .task-status.progress { background: #3b82f6; }
        .task-status.pending { background: #6b7280; }
        
        .task-dates {
            font-size: 12px;
            color: #6b7280;
            margin-bottom: 6px;
        }
        .task-note {
            font-size: 12px;
            color: #9ca3af;
            line-height: 1.4;
        }
        
        .task.overdue {
            border: 2px solid #ef4444;
            background: #fef2f2;
        }
        .task.warning-soon {
            border: 2px solid #f59e0b;
            background: #fffbeb;
        }
        .task.should-start {
            border: 2px solid #f59e0b;
            background: #fffbeb;
        }
        .date-warning {
            font-size: 11px;
            color: #dc2626;
            margin-top: 4px;
            font-weight: 500;
        }
        
        .milestones-section {
            background: white;
            border-radius: 16px;
            padding: 24px;
            margin-top: 20px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }
        .milestones-section h2 {
            font-size: 20px;
            color: #1f2937;
            margin-bottom: 16px;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        .milestones-grid {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(300px, 1fr));
            gap: 12px;
        }
        .milestone {
            background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%);
            border-radius: 8px;
            padding: 12px;
            border-left: 4px solid #f59e0b;
        }
        .milestone.completed {
            background: linear-gradient(135deg, #d1fae5 0%, #a7f3d0 100%);
            border-left-color: #10b981;
        }
        .milestone-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 4px;
        }
        .milestone-name {
            font-size: 14px;
            font-weight: 500;
            color: #1f2937;
        }
        .milestone-dates {
            font-size: 12px;
            color: #6b7280;
        }
        
        .footer {
            text-align: center;
            padding: 20px;
            color: rgba(255,255,255,0.8);
            font-size: 14px;
        }
        
        .filter-control {
            display: flex;
            align-items: center;
            gap: 16px;
            margin-top: 16px;
            padding-top: 16px;
            border-top: 1px solid #e5e7eb;
        }
        .filter-label {
            font-size: 14px;
            color: #4b5563;
            font-weight: 500;
        }
        .filter-checkboxes {
            display: flex;
            align-items: center;
            gap: 12px;
        }
        .checkbox-item {
            display: flex;
            align-items: center;
            gap: 6px;
            cursor: pointer;
        }
        .checkbox-item input[type="checkbox"] {
            width: 16px;
            height: 16px;
            cursor: pointer;
            accent-color: #3b82f6;
        }
        .checkbox-item.completed input[type="checkbox"] {
            accent-color: #10b981;
        }
        .checkbox-item.pending input[type="checkbox"] {
            accent-color: #6b7280;
        }
        .checkbox-item label {
            font-size: 14px;
            color: #4b5563;
            cursor: pointer;
            user-select: none;
        }
        .checkbox-item.completed label { color: #10b981; }
        .checkbox-item.progress label { color: #3b82f6; }
        .checkbox-item.pending label { color: #6b7280; }
        .task.hidden {
            display: none;
        }
    </style>
'''

JS_CODE = '''
    <script>
        function filterTasks() {
            const showCompleted = document.getElementById('showCompleted').checked;
            const showProgress = document.getElementById('showProgress').checked;
            const showPending = document.getElementById('showPending').checked;
            
            document.querySelectorAll('.task').forEach(el => {
                const isCompleted = el.classList.contains('completed');
                const isProgress = el.classList.contains('progress');
                const isPending = el.classList.contains('pending');
                
                const shouldShow = (isCompleted && showCompleted) || 
                                   (isProgress && showProgress) || 
                                   (isPending && showPending);
                
                el.classList.toggle('hidden', !shouldShow);
            });
        }
        
        document.addEventListener('DOMContentLoaded', function() {
            filterTasks();
        });
    </script>
'''

def generate_html(categories, milestones):
    total_tasks = sum(len(c['tasks']) for c in categories)
    completed = sum(1 for c in categories for t in c['tasks'] if t['status'] == '已完成')
    in_progress = sum(1 for c in categories for t in c['tasks'] if t['status'] == '进行中')
    
    progress_pct = int(completed / total_tasks * 100) if total_tasks > 0 else 0
    
    date_warnings = []
    for c in categories:
        for t in c['tasks']:
            date_status, date_msg = calculate_date_status(t['start_date'], t['end_date'], t['status'])
            if date_status in ['overdue', 'warning', 'soon', 'should_start']:
                date_warnings.append({
                    'task': t['name'],
                    'category': c['name'],
                    'status': date_status,
                    'message': date_msg
                })
    
    warnings_display = 'block' if date_warnings else 'none'
    warnings_html = ''.join([f'<div class="warning-item"><span class="warning-badge{" orange" if w["status"] == "should_start" else ""}">{w["message"]}</span><span>{w["task"]}</span><span style="color:#6b7280">({w["category"]})</span></div>' for w in date_warnings[:5]])
    
    html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>模型推理服务工作看板</title>
{CSS_STYLES}
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="header-top">
                <h1>📊 模型推理服务工作看板</h1>
                <div class="header-date">📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
            </div>
            
            <div class="progress-section">
                <div class="progress-label">
                    <span>整体进度</span>
                    <span>{completed}/{total_tasks} 已完成</span>
                </div>
                <div class="progress-bar">
                    <div class="progress-fill" style="width: {progress_pct}%">{progress_pct}%</div>
                </div>
            </div>
            
            <div class="stats">
                <div class="stat-card completed">
                    <div class="stat-value">{completed}</div>
                    <div class="stat-label">已完成</div>
                </div>
                <div class="stat-card progress">
                    <div class="stat-value">{in_progress}</div>
                    <div class="stat-label">进行中</div>
                </div>
                <div class="stat-card pending">
                    <div class="stat-value">{total_tasks - completed - in_progress}</div>
                    <div class="stat-label">未启动</div>
                </div>
                <div class="stat-card warning">
                    <div class="stat-value">{len(date_warnings)}</div>
                    <div class="stat-label">需关注</div>
                </div>
            </div>
            
            <div class="warnings" style="display:{warnings_display}">
                <h3>⚠️ 日期提醒</h3>
                {warnings_html}
            </div>
            
            <div class="filter-control">
                <span class="filter-label">显示状态：</span>
                <div class="filter-checkboxes">
                    <div class="checkbox-item progress">
                        <input type="checkbox" id="showProgress" checked onchange="filterTasks()">
                        <label for="showProgress">进行中</label>
                    </div>
                    <div class="checkbox-item completed">
                        <input type="checkbox" id="showCompleted" onchange="filterTasks()">
                        <label for="showCompleted">已完成</label>
                    </div>
                    <div class="checkbox-item pending">
                        <input type="checkbox" id="showPending" onchange="filterTasks()">
                        <label for="showPending">未启动</label>
                    </div>
                </div>
            </div>
        </div>
        
        <div class="kanban">
'''
    
    for category in categories:
        cat_tasks = category['tasks']
        cat_completed = sum(1 for t in cat_tasks if t['status'] == '已完成')
        cat_total = len(cat_tasks)
        cat_pct = int(cat_completed / cat_total * 100) if cat_total > 0 else 0
        
        html += f'''
            <div class="category">
                <div class="category-header">
                    <div class="category-title">{category['name']}</div>
                    <div class="category-progress">
                        <span>{cat_completed}/{cat_total}</span>
                        <div class="category-progress-bar">
                            <div class="category-progress-fill" style="width: {cat_pct}%"></div>
                        </div>
                        <span>{cat_pct}%</span>
                    </div>
                </div>
                <div class="tasks">
'''
        
        for task in cat_tasks:
            status_class = task['status'].replace(' ', '-')
            if task['status'] == '进行中':
                status_class = 'progress'
            elif task['status'] == '未启动':
                status_class = 'pending'
            elif task['status'] == '已完成':
                status_class = 'completed'
            
            date_status, date_msg = calculate_date_status(task['start_date'], task['end_date'], task['status'])
            task_class = f'task {status_class}'
            if date_status == 'overdue':
                task_class += ' overdue'
            elif date_status in ['warning', 'soon']:
                task_class += ' warning-soon'
            elif date_status == 'should_start':
                task_class += ' should-start'
            
            date_range = f"{format_date(task['start_date'])} - {format_date(task['end_date'])}"
            
            date_warning_html = f'<div class="date-warning">⚠️ {date_msg}</div>' if date_status in ['overdue', 'warning', 'should_start'] else ''
            note_html = f'<div class="task-note">{task["note"][:80]}{"..." if len(str(task["note"])) > 80 else ""}</div>' if task['note'] else ''
            
            html += f'''
                    <div class="{task_class}">
                        <div class="task-header">
                            <div class="task-name">#<span class="task-row">{task['id']}</span> {task['name']}</div>
                            <div class="task-status {status_class}">{task['status']}</div>
                        </div>
                        <div class="task-dates">📅 {date_range}</div>
                        {date_warning_html}
                        {note_html}
                    </div>
'''
        
        html += '''
                </div>
            </div>
'''
    
    html += '''
        </div>
'''
    
    if milestones:
        html += '''
        <div class="milestones-section">
            <h2>🎯 里程碑</h2>
            <div class="milestones-grid">
'''
        for m in milestones:
            m_status_class = 'completed' if m['status'] == '已完成' else ''
            date_range = f"{format_date(m['start_date'])} - {format_date(m['end_date'])}"
            m_status_html = 'completed' if m['status'] == '已完成' else 'pending'
            deliverable_html = f'<div class="task-note">📦 {m["deliverable"]}</div>' if m['deliverable'] else ''
            
            html += f'''
                <div class="milestone {m_status_class}">
                    <div class="milestone-header">
                        <div class="milestone-name">#<span class="task-row">{m['id']}</span> {m['name']}</div>
                        <div class="task-status {m_status_html}">{m['status']}</div>
                    </div>
                    <div class="milestone-dates">📅 {date_range}</div>
                    {deliverable_html}
                </div>
'''
        html += '''
            </div>
        </div>
'''
    
    html += f'''
        <div class="footer">
            📋 数据来源: 模型推理服务.xlsx | 手动刷新页面查看最新数据
        </div>
    </div>
{JS_CODE}
</body>
</html>
'''
    
    return html

def main():
    categories, milestones = read_excel_data()
    html = generate_html(categories, milestones)
    
    with open(HTML_FILE, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f'[OK] HTML generated: {HTML_FILE}')
    print(f'[INFO] Tasks: {sum(len(c["tasks"]) for c in categories)}, Milestones: {len(milestones)}')

if __name__ == '__main__':
    main()