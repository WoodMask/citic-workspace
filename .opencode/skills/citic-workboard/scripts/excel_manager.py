# -*- coding: utf-8 -*-
"""
Excel Manager for Model Inference Service Workboard
Provides CRUD operations and utilities for managing work items.

File paths use relative paths based on workspace directory.
"""

import os
import sys
import subprocess
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Any

SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def get_workspace_dir() -> str:
    """
    Get workspace directory from skill directory.
    Skill is located at: {workspace}/.opencode/skills/citic-workboard/
    So workspace is: {skill_dir}/../../
    """
    return os.path.dirname(os.path.dirname(os.path.dirname(SKILL_DIR)))


EXCEL_FILE = None
WORKBOARD_SCRIPT = None


def init_paths(workspace_dir: str = None):
    """
    Initialize file paths based on workspace directory.
    
    Args:
        workspace_dir: The workspace directory path. If None, derives from skill location.
    
    Returns:
        Excel file path.
    """
    if workspace_dir is None:
        workspace_dir = get_workspace_dir()
    
    global EXCEL_FILE, WORKBOARD_SCRIPT
    EXCEL_FILE = os.path.join(workspace_dir, '模型推理服务.xlsx')
    WORKBOARD_SCRIPT = os.path.join(workspace_dir, 'workboard.py')
    
    return EXCEL_FILE


def get_categories() -> List[str]:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    categories = []
    for row_idx in range(2, ws.max_row + 1):
        task_name = ws.cell(row=row_idx, column=2).value
        if task_name and str(task_name).strip().startswith(('一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '里程碑')):
            categories.append(str(task_name).strip())
    
    wb.close()
    return categories


def find_category_row(category_name: str) -> Optional[int]:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    category_name = category_name.strip()
    
    for row_idx in range(2, ws.max_row + 1):
        task_name = ws.cell(row=row_idx, column=2).value
        if task_name:
            task_str = str(task_name).strip()
            if task_str.startswith(('一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '里程碑')):
                if category_name in task_str or task_str.startswith(category_name):
                    wb.close()
                    return row_idx
    
    wb.close()
    return None


def get_last_task_row_in_category(category_row: int) -> int:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    last_row = category_row
    for row_idx in range(category_row + 1, ws.max_row + 1):
        task_name = ws.cell(row=row_idx, column=2).value
        if task_name:
            task_str = str(task_name).strip()
            if task_str.startswith(('一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '里程碑')):
                break
            status = ws.cell(row=row_idx, column=3).value
            if status and str(status).strip() in ['已完成', '进行中', '未启动']:
                last_row = row_idx
    
    wb.close()
    return last_row


def get_task_by_row(row: int) -> Optional[Dict[str, Any]]:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    task_id = ws.cell(row=row, column=1).value
    task_name = ws.cell(row=row, column=2).value
    status = ws.cell(row=row, column=3).value
    start_date = ws.cell(row=row, column=4).value
    end_date = ws.cell(row=row, column=5).value
    deliverable = ws.cell(row=row, column=6).value
    note = ws.cell(row=row, column=7).value
    
    wb.close()
    
    if task_name is None:
        return None
    
    return {
        'row': row,
        'id': str(task_id).strip() if task_id else '',
        'name': str(task_name).strip() if task_name else '',
        'status': str(status).strip() if status else '',
        'start_date': start_date,
        'end_date': end_date,
        'deliverable': str(deliverable).strip() if deliverable else '',
        'note': str(note).strip() if note else '',
    }


def get_tasks_by_category(category: str) -> List[Dict[str, Any]]:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    tasks = []
    current_category = None
    in_target_category = False
    
    for row_idx in range(2, ws.max_row + 1):
        task_name = ws.cell(row=row_idx, column=2).value
        status = ws.cell(row=row_idx, column=3).value
        
        if task_name:
            task_str = str(task_name).strip()
            
            if task_str.startswith(('一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '里程碑')):
                current_category = task_str
                in_target_category = (category in task_str or task_str.startswith(category))
            elif in_target_category and status and str(status).strip() in ['已完成', '进行中', '未启动']:
                tasks.append(get_task_by_row(row_idx))
    
    wb.close()
    return tasks


def get_tasks_by_status(status: str) -> List[Dict[str, Any]]:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    tasks = []
    
    for row_idx in range(2, ws.max_row + 1):
        task_status = ws.cell(row=row_idx, column=3).value
        if task_status and str(task_status).strip() == status:
            tasks.append(get_task_by_row(row_idx))
    
    wb.close()
    return tasks


def get_max_task_id() -> int:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    max_id = 0
    for row_idx in range(2, ws.max_row + 1):
        task_id = ws.cell(row=row_idx, column=1).value
        if task_id:
            try:
                max_id = max(max_id, int(str(task_id).strip()))
            except:
                pass
    
    wb.close()
    return max_id

def get_task_by_id(task_id: str) -> Optional[Dict[str, Any]]:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    task_id = str(task_id).strip().lstrip('#')
    try:
        task_id = f'{int(task_id):03d}'
    except ValueError:
        wb.close()
        return None
    
    for row_idx in range(2, ws.max_row + 1):
        cell_id = ws.cell(row=row_idx, column=1).value
        if cell_id and str(cell_id).strip() == task_id:
            wb.close()
            return get_task_by_row(row_idx)
    
    wb.close()
    return None

def find_row_by_id(task_id: str) -> Optional[int]:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    task_id = str(task_id).strip().lstrip('#')
    try:
        task_id = f'{int(task_id):03d}'
    except ValueError:
        wb.close()
        return None
    
    for row_idx in range(2, ws.max_row + 1):
        cell_id = ws.cell(row=row_idx, column=1).value
        if cell_id and str(cell_id).strip() == task_id:
            wb.close()
            return row_idx
    
    wb.close()
    return None

def add_task(
    name: str,
    category: str,
    status: str = '未启动',
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    note: Optional[str] = None
) -> Dict[str, Any]:
    if EXCEL_FILE is None:
        init_paths()
    
    category_row = find_category_row(category)
    if category_row is None:
        raise ValueError(f"Category '{category}' not found")
    
    last_row = get_last_task_row_in_category(category_row)
    insert_row = last_row + 1
    
    new_id = get_max_task_id() + 1
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    ws.insert_rows(insert_row)
    
    ws.cell(row=insert_row, column=1).value = f'{new_id:03d}'
    ws.cell(row=insert_row, column=2).value = name
    ws.cell(row=insert_row, column=3).value = status
    ws.cell(row=insert_row, column=4).value = start_date
    ws.cell(row=insert_row, column=5).value = end_date
    ws.cell(row=insert_row, column=6).value = ''
    ws.cell(row=insert_row, column=7).value = note if note else ''
    
    wb.save(EXCEL_FILE)
    wb.close()
    
    return {
        'row': insert_row,
        'id': f'{new_id:03d}',
        'name': name,
        'status': status,
        'start_date': start_date,
        'end_date': end_date,
        'note': note,
        'category': category
    }


def update_task_status(row: int, status: str) -> bool:
    if EXCEL_FILE is None:
        init_paths()
    
    if status not in ['已完成', '进行中', '未启动']:
        raise ValueError(f"Invalid status: {status}. Must be 已完成/进行中/未启动")
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    current_status = ws.cell(row=row, column=3).value
    if current_status is None or str(current_status).strip() not in ['已完成', '进行中', '未启动']:
        wb.close()
        return False
    
    ws.cell(row=row, column=3).value = status
    wb.save(EXCEL_FILE)
    wb.close()
    
    return True

def update_task_status_by_id(task_id: str, status: str) -> bool:
    row = find_row_by_id(task_id)
    if row is None:
        return False
    return update_task_status(row, status)


def update_task_dates(
    row: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> bool:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    current_status = ws.cell(row=row, column=3).value
    if current_status is None or str(current_status).strip() not in ['已完成', '进行中', '未启动']:
        wb.close()
        return False
    
    if start_date is not None:
        ws.cell(row=row, column=4).value = start_date
    if end_date is not None:
        ws.cell(row=row, column=5).value = end_date
    
    wb.save(EXCEL_FILE)
    wb.close()
    
    return True

def update_task_dates_by_id(task_id: str, start_date: Optional[date] = None, end_date: Optional[date] = None) -> bool:
    row = find_row_by_id(task_id)
    if row is None:
        return False
    return update_task_dates(row, start_date, end_date)


def clear_task_dates(row: int) -> bool:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    current_status = ws.cell(row=row, column=3).value
    if current_status is None or str(current_status).strip() not in ['已完成', '进行中', '未启动']:
        wb.close()
        return False
    
    ws.cell(row=row, column=4).value = None
    ws.cell(row=row, column=5).value = None
    
    wb.save(EXCEL_FILE)
    wb.close()
    
    return True

def clear_task_dates_by_id(task_id: str) -> bool:
    row = find_row_by_id(task_id)
    if row is None:
        return False
    return clear_task_dates(row)


def update_task_note(row: int, note: str, append: bool = False) -> bool:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    current_status = ws.cell(row=row, column=3).value
    if current_status is None or str(current_status).strip() not in ['已完成', '进行中', '未启动']:
        wb.close()
        return False
    
    current_note = ws.cell(row=row, column=7).value
    
    if append and current_note:
        ws.cell(row=row, column=7).value = f"{current_note}; {note}"
    else:
        ws.cell(row=row, column=7).value = note
    
    wb.save(EXCEL_FILE)
    wb.close()
    
    return True

def update_task_note_by_id(task_id: str, note: str, append: bool = False) -> bool:
    row = find_row_by_id(task_id)
    if row is None:
        return False
    return update_task_note(row, note, append)


def delete_task(row: int) -> bool:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    current_status = ws.cell(row=row, column=3).value
    if current_status is None or str(current_status).strip() not in ['已完成', '进行中', '未启动']:
        wb.close()
        return False
    
    ws.delete_rows(row)
    wb.save(EXCEL_FILE)
    wb.close()
    
    return True

def delete_task_by_id(task_id: str) -> bool:
    row = find_row_by_id(task_id)
    if row is None:
        return False
    return delete_task(row)


def parse_relative_date(date_str: str) -> Optional[date]:
    """
    Parse relative date string to actual date.
    
    Args:
        date_str: Date string (YYYY-MM-DD, MM/DD, today, tomorrow, next week, etc.)
    
    Returns:
        date object or None if invalid.
    """
    today = date.today()
    
    date_str = date_str.strip().lower()
    
    if date_str in ['today', '今天', '今日']:
        return today
    elif date_str in ['tomorrow', '明天', '明日']:
        return today + timedelta(days=1)
    elif date_str in ['后天', '后日']:
        return today + timedelta(days=2)
    elif date_str in ['next week', '下周']:
        days_until_monday = (7 - today.weekday()) % 7
        if days_until_monday == 0:
            days_until_monday = 7
        return today + timedelta(days=days_until_monday)
    elif date_str in ['下周一']:
        days_until_monday = (0 - today.weekday()) % 7
        if days_until_monday == 0:
            days_until_monday = 7
        return today + timedelta(days=days_until_monday)
    elif date_str in ['月底', '月末', 'end of month']:
        if today.month == 12:
            return date(today.year + 1, 1, 1) - timedelta(days=1)
        else:
            return date(today.year, today.month + 1, 1) - timedelta(days=1)
    
    try:
        if len(date_str) == 10 and '-' in date_str:
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        elif '/' in date_str:
            parts = date_str.split('/')
            if len(parts) == 2:
                month, day = int(parts[0]), int(parts[1])
                return date(today.year, month, day)
            elif len(parts) == 3:
                year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
                return date(year, month, day)
    except:
        pass
    
    return None


def generate_progress_report() -> Dict[str, Any]:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    stats = {
        'total': 0,
        'completed': 0,
        'in_progress': 0,
        'not_started': 0,
        'categories': {},
        'milestones': {'total': 0, 'completed': 0, 'in_progress': 0, 'not_started': 0}
    }
    
    current_category = None
    is_milestone = False
    
    for row_idx in range(2, ws.max_row + 1):
        task_name = ws.cell(row=row_idx, column=2).value
        status = ws.cell(row=row_idx, column=3).value
        
        if task_name:
            task_str = str(task_name).strip()
            
            if task_str.startswith(('一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '里程碑')):
                current_category = task_str
                is_milestone = task_str == '里程碑'
                if not is_milestone and current_category not in stats['categories']:
                    stats['categories'][current_category] = {'total': 0, 'completed': 0, 'in_progress': 0, 'not_started': 0}
            elif status and str(status).strip() in ['已完成', '进行中', '未启动']:
                status_str = str(status).strip()
                stats['total'] += 1
                
                if status_str == '已完成':
                    stats['completed'] += 1
                elif status_str == '进行中':
                    stats['in_progress'] += 1
                else:
                    stats['not_started'] += 1
                
                if is_milestone:
                    stats['milestones']['total'] += 1
                    if status_str == '已完成':
                        stats['milestones']['completed'] += 1
                    elif status_str == '进行中':
                        stats['milestones']['in_progress'] += 1
                    else:
                        stats['milestones']['not_started'] += 1
                elif current_category:
                    stats['categories'][current_category]['total'] += 1
                    if status_str == '已完成':
                        stats['categories'][current_category]['completed'] += 1
                    elif status_str == '进行中':
                        stats['categories'][current_category]['in_progress'] += 1
                    else:
                        stats['categories'][current_category]['not_started'] += 1
    
    wb.close()
    
    stats['progress_pct'] = int(stats['completed'] / stats['total'] * 100) if stats['total'] > 0 else 0
    
    return stats


def get_overdue_tasks(days_threshold: int = 7) -> List[Dict[str, Any]]:
    if EXCEL_FILE is None:
        init_paths()
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    today = date.today()
    overdue_tasks = []
    
    current_category = None
    
    for row_idx in range(2, ws.max_row + 1):
        task_id = ws.cell(row=row_idx, column=1).value
        task_name = ws.cell(row=row_idx, column=2).value
        status = ws.cell(row=row_idx, column=3).value
        start_date_cell = ws.cell(row=row_idx, column=4).value
        end_date_cell = ws.cell(row=row_idx, column=5).value
        
        if task_name:
            task_str = str(task_name).strip()
            
            if task_str.startswith(('一、', '二、', '三、', '四、', '五、', '六、', '七、', '八、', '里程碑')):
                current_category = task_str
            elif status and str(status).strip() in ['已完成', '进行中', '未启动']:
                status_str = str(status).strip()
                
                if status_str == '已完成':
                    continue
                
                start_date = None
                if isinstance(start_date_cell, datetime):
                    start_date = start_date_cell.date()
                elif isinstance(start_date_cell, date):
                    start_date = start_date_cell
                
                if status_str == '未启动' and start_date and start_date <= today:
                    days_passed = (today - start_date).days
                    msg = f"应启动{days_passed}天" if days_passed > 0 else "今日应启动"
                    overdue_tasks.append({
                        'row': row_idx,
                        'id': str(task_id).strip() if task_id else '',
                        'name': task_str,
                        'status': status_str,
                        'start_date': start_date,
                        'days_passed': days_passed,
                        'alert_type': 'should_start',
                        'message': msg,
                        'category': current_category
                    })
                
                if end_date_cell:
                    end_date = None
                    if isinstance(end_date_cell, datetime):
                        end_date = end_date_cell.date()
                    elif isinstance(end_date_cell, date):
                        end_date = end_date_cell
                    
                    if end_date:
                        days_diff = (end_date - today).days
                        
                        if days_diff < 0:
                            overdue_tasks.append({
                                'row': row_idx,
                                'id': str(task_id).strip() if task_id else '',
                                'name': task_str,
                                'status': status_str,
                                'end_date': end_date,
                                'days_diff': days_diff,
                                'alert_type': 'overdue',
                                'message': f"超期{abs(days_diff)}天",
                                'category': current_category
                            })
                        elif days_diff <= days_threshold and status_str == '进行中':
                            overdue_tasks.append({
                                'row': row_idx,
                                'id': str(task_id).strip() if task_id else '',
                                'name': task_str,
                                'status': status_str,
                                'end_date': end_date,
                                'days_diff': days_diff,
                                'alert_type': 'due_soon',
                                'message': f"{days_diff}天后到期",
                                'category': current_category
                            })
    
    wb.close()
    return overdue_tasks


def refresh_board(workspace_dir: str = None) -> bool:
    """
    Refresh HTML board by running workboard.py.
    
    Args:
        workspace_dir: Workspace directory path.
    
    Returns:
        True if successful, False otherwise.
    """
    if workspace_dir is None:
        workspace_dir = get_workspace_dir()
    
    workboard_path = os.path.join(workspace_dir, 'workboard.py')
    
    if not os.path.exists(workboard_path):
        print(f"[ERROR] workboard.py not found at {workboard_path}")
        return False
    
    try:
        result = subprocess.run(
            [sys.executable, workboard_path],
            cwd=workspace_dir,
            capture_output=True,
            text=True
        )
        
        if result.returncode == 0:
            print("[OK] HTML board refreshed")
            return True
        else:
            print(f"[ERROR] Failed to refresh board: {result.stderr}")
            return False
    except Exception as e:
        print(f"[ERROR] Exception: {e}")
        return False


def format_progress_report(report: Dict[str, Any]) -> str:
    """
    Format progress report as readable text.
    
    Args:
        report: Progress report dictionary.
    
    Returns:
        Formatted text string.
    """
    lines = []
    lines.append("=" * 50)
    lines.append("Progress Report - Model Inference Service")
    lines.append("=" * 50)
    lines.append("")
    
    lines.append("Overall Progress:")
    lines.append(f"   Total: {report['total']}")
    lines.append(f"   Completed: {report['completed']} ({report['progress_pct']}%)")
    lines.append(f"   In Progress: {report['in_progress']}")
    lines.append(f"   Not Started: {report['not_started']}")
    lines.append("")
    
    lines.append("Category Progress:")
    for cat, cat_stats in report['categories'].items():
        cat_pct = int(cat_stats['completed'] / cat_stats['total'] * 100) if cat_stats['total'] > 0 else 0
        lines.append(f"   {cat}: {cat_stats['completed']}/{cat_stats['total']} ({cat_pct}%)")
    lines.append("")
    
    lines.append("Milestones:")
    ms = report['milestones']
    ms_pct = int(ms['completed'] / ms['total'] * 100) if ms['total'] > 0 else 0
    lines.append(f"   Completed: {ms['completed']}/{ms['total']} ({ms_pct}%)")
    lines.append("")
    
    return "\n".join(lines)


def format_task_info(task: Dict[str, Any]) -> str:
    lines = []
    lines.append(f"Task #{task.get('id', task['row'])}")
    lines.append(f"   Name: {task['name']}")
    lines.append(f"   Status: {task['status']}")
    
    start = task['start_date']
    end = task['end_date']
    
    if start:
        if isinstance(start, (datetime, date)):
            start_str = start.strftime('%Y-%m-%d')
        else:
            start_str = str(start)
    else:
        start_str = 'Not set'
    
    if end:
        if isinstance(end, (datetime, date)):
            end_str = end.strftime('%Y-%m-%d')
        else:
            end_str = str(end)
    else:
        end_str = 'Not set'
    
    lines.append(f"   Start: {start_str}")
    lines.append(f"   End: {end_str}")
    
    if task['note']:
        lines.append(f"   Note: {task['note']}")
    
    return "\n".join(lines)


def format_tasks_list(tasks: List[Dict[str, Any]], title: str = "Tasks") -> str:
    lines = []
    lines.append(f"{title} ({len(tasks)} items)")
    lines.append("-" * 40)
    
    for task in tasks:
        status_icon = '[DONE]' if task['status'] == '已完成' else ('[PROG]' if task['status'] == '进行中' else '[TODO]')
        task_id = task.get('id', task['row'])
        lines.append(f"#{task_id} {status_icon} {task['name']}")
    
    return "\n".join(lines)


try:
    from openpyxl import load_workbook
except ImportError:
    print("[WARN] openpyxl not installed, Excel operations will fail")


if __name__ == '__main__':
    init_paths()
    
    report = generate_progress_report()
    print(format_progress_report(report))
    
    overdue = get_overdue_tasks()
    if overdue:
        print("\n[ALERT] Tasks need attention:")
        for t in overdue:
            print(f"   #{t['row']} {t['name']} - {t['message']}")